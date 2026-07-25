"""Cash Flow Intelligence module — Sprint 5, Day 31-32.
Builds output/cashflow_intelligence.xlsx, output/distress_alerts.csv, output/pattern_changes.csv.
"""
import os
import csv
import sqlite3
import pandas as pd

from src.analytics.cashflow_kpis import cfo_quality_score, capex_intensity, fcf_conversion_rate
from src.analytics.cagr import compute_cagr

DB_PATH = os.getenv("DB_PATH", "nifty100.db")


def _distress_flag(cfo_latest, cff_latest):
    return pd.notna(cfo_latest) and pd.notna(cff_latest) and cfo_latest < 0 and cff_latest > 0


def _deleveraging_flag(cff_latest, borrowings_series):
    if pd.isna(cff_latest) or cff_latest >= 0:
        return False
    vals = borrowings_series.dropna().tolist()
    if len(vals) < 2:
        return False
    return vals[-1] < vals[-2]


def build_cashflow_intelligence(conn):
    cf = pd.read_sql_query("SELECT * FROM cashflow ORDER BY company_id, year", conn)
    pl = pd.read_sql_query("SELECT company_id, year, sales, operating_profit, net_profit FROM profitandloss", conn)
    bs = pd.read_sql_query("SELECT company_id, year, borrowings FROM balancesheet", conn)
    sec = pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", conn)
    capital_alloc_path = "output/capital_allocation.csv"
    capital_alloc = pd.read_csv(capital_alloc_path) if os.path.exists(capital_alloc_path) else pd.DataFrame()

    rows = []
    for cid, cf_hist in cf.groupby("company_id"):
        cf_hist = cf_hist.sort_values("year")
        pl_hist = pl[pl["company_id"] == cid].sort_values("year")
        bs_hist = bs[bs["company_id"] == cid].sort_values("year")
        sector = sec.loc[sec["company_id"] == cid, "broad_sector"].iloc[0] if cid in sec["company_id"].values else None

        merged = cf_hist.merge(pl_hist, on=["company_id", "year"], how="left").merge(
            bs_hist, on=["company_id", "year"], how="left")
        if merged.empty:
            continue

        cfo_score, cfo_label = cfo_quality_score(merged["operating_activity"].tolist(),
                                                  merged["net_profit"].tolist())
        latest = merged.iloc[-1]
        capex_pct, capex_label = capex_intensity(latest.get("investing_activity"), latest.get("sales"))

        fcf_series = (merged["operating_activity"].fillna(0) + merged["investing_activity"].fillna(0))
        fcf_start, fcf_end = (fcf_series.iloc[0], fcf_series.iloc[-1]) if len(fcf_series) >= 5 else (None, None)
        fcf_cagr = None
        if len(fcf_series) >= 5:
            fcf_cagr, _flag = compute_cagr(fcf_series.iloc[-5], fcf_series.iloc[-1], 5)

        fcf_latest = fcf_series.iloc[-1] if not fcf_series.empty else None
        fcf_conv = fcf_conversion_rate(fcf_latest, latest.get("operating_profit"))

        distress = _distress_flag(latest.get("operating_activity"), latest.get("financing_activity"))
        deleveraging = _deleveraging_flag(latest.get("financing_activity"), merged["borrowings"])

        pattern_label = None
        if not capital_alloc.empty:
            match = capital_alloc[(capital_alloc["company_id"] == cid)].sort_values("year")
            if not match.empty:
                pattern_label = match.iloc[-1]["pattern_label"]

        rows.append({
            "company_id": cid, "sector": sector, "cfo_quality_score": round(cfo_score, 2) if cfo_score else None,
            "cfo_quality_label": cfo_label, "capex_intensity_pct": round(capex_pct, 2) if capex_pct is not None else None,
            "capex_label": capex_label, "fcf_cagr_5yr": round(fcf_cagr, 2) if fcf_cagr is not None else None,
            "fcf_conversion_pct": round(fcf_conv, 2) if fcf_conv is not None else None,
            "distress_flag": distress, "deleveraging_flag": deleveraging,
            "capital_allocation_label": pattern_label,
            "_cfo_latest": latest.get("operating_activity"), "_cff_latest": latest.get("financing_activity"),
            "_net_profit_latest": latest.get("net_profit"),
        })

    return pd.DataFrame(rows)


def build_pattern_changes(capital_alloc_path="output/capital_allocation.csv"):
    if not os.path.exists(capital_alloc_path):
        return pd.DataFrame()
    df = pd.read_csv(capital_alloc_path).sort_values(["company_id", "year"])
    changes = []
    for cid, hist in df.groupby("company_id"):
        hist = hist.reset_index(drop=True)
        for i in range(1, len(hist)):
            prev, curr = hist.iloc[i - 1], hist.iloc[i]
            if prev["pattern_label"] != curr["pattern_label"]:
                changes.append({"company_id": cid, "from_year": prev["year"], "to_year": curr["year"],
                                "from_pattern": prev["pattern_label"], "to_pattern": curr["pattern_label"]})
    return pd.DataFrame(changes)


def export_cashflow_intelligence(df, path="output/cashflow_intelligence.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    os.makedirs("output", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Intelligence"
    display_cols = [c for c in df.columns if not c.startswith("_")]
    ws.append(display_cols)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    for _, row in df.iterrows():
        ws.append([row[c] for c in display_cols])
    for r_row in ws.iter_rows(min_row=2):
        for cell in r_row:
            cell.font = Font(name="Arial")
    wb.save(path)
    return path


def run(db_path=None):
    conn = sqlite3.connect(db_path or DB_PATH)
    df = build_cashflow_intelligence(conn)
    if df.empty:
        print("No cashflow data available.")
        conn.close()
        return df

    export_cashflow_intelligence(df)

    distress = df[df["distress_flag"] == True][["company_id", "_cfo_latest", "_cff_latest", "_net_profit_latest"]] \
        .rename(columns={"_cfo_latest": "cfo_value", "_cff_latest": "cff_value", "_net_profit_latest": "latest_net_profit"})
    distress.to_csv("output/distress_alerts.csv", index=False)

    pattern_changes = build_pattern_changes()
    pattern_changes.to_csv("output/pattern_changes.csv", index=False)

    print(f"cashflow_intelligence.xlsx: {len(df)} rows. distress_alerts.csv: {len(distress)} flagged. "
          f"pattern_changes.csv: {len(pattern_changes)} changes.")
    conn.close()
    return df


if __name__ == "__main__":
    run()
