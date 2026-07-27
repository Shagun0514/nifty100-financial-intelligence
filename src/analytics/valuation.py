"""Valuation module — Sprint 4, Day 26.
Uses market_cap + financial_ratios + sectors to compute FCF yield and
overvaluation/discount flags, writes valuation_summary.xlsx and valuation_flags.csv.
"""
import os
import sqlite3
import pandas as pd

DB_PATH = os.getenv("DB_PATH", "nifty100.db")


def fcf_yield_pct(fcf_cr, market_cap_crore):
    """FCF / market_cap * 100. None if market_cap missing/zero."""
    if not market_cap_crore or pd.isna(market_cap_crore) or fcf_cr is None or pd.isna(fcf_cr):
        return None
    return fcf_cr / market_cap_crore * 100


def classify_valuation(pe, sector_median_pe):
    """Caution if PE > 1.5x sector median; Discount if PE < 0.7x; else Fair.
    Returns 'Fair' (not flagged) if either input is missing — can't judge without both."""
    if pe is None or pd.isna(pe) or sector_median_pe is None or pd.isna(sector_median_pe) or sector_median_pe == 0:
        return "Fair"
    if pe > sector_median_pe * 1.5:
        return "Caution"
    if pe < sector_median_pe * 0.7:
        return "Discount"
    return "Fair"


def build_valuation_table(conn):
    comp = pd.read_sql_query("SELECT id as company_id, company_name from companies", conn)
    sec = pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", conn)
    mc = pd.read_sql_query(
        "SELECT company_id, year, pe_ratio, pb_ratio, ev_ebitda, market_cap_crore FROM market_cap", conn)
    fr = pd.read_sql_query("SELECT company_id, year, free_cash_flow_cr FROM financial_ratios", conn)

    if mc.empty:
        return pd.DataFrame()

    mc_latest = mc.sort_values("year").groupby("company_id").tail(1)
    fr_latest = fr.sort_values("year").groupby("company_id").tail(1)[["company_id", "free_cash_flow_cr"]]

    df = comp.merge(sec, on="company_id", how="left").merge(mc_latest, on="company_id", how="left")
    df = df.merge(fr_latest, on="company_id", how="left")

    # sector median P/E (latest year, across all companies in that sector with mc data)
    sector_median = mc_latest.merge(sec, on="company_id", how="left").groupby("broad_sector")["pe_ratio"].median()
    df["sector_median_pe"] = df["broad_sector"].map(sector_median)

    # 5-year median P/E per company (from full history, not just latest)
    five_yr_median = mc.sort_values("year").groupby("company_id").tail(5).groupby("company_id")["pe_ratio"].median()
    df["five_yr_median_pe"] = df["company_id"].map(five_yr_median)

    df["fcf_yield_pct"] = df.apply(lambda r: fcf_yield_pct(r["free_cash_flow_cr"], r["market_cap_crore"]), axis=1)
    df["pe_vs_sector_median_pct"] = df.apply(
        lambda r: None if pd.isna(r["pe_ratio"]) or pd.isna(r["sector_median_pe"]) or r["sector_median_pe"] == 0
        else (r["pe_ratio"] / r["sector_median_pe"] - 1) * 100, axis=1)
    df["flag"] = df.apply(lambda r: classify_valuation(r["pe_ratio"], r["sector_median_pe"]), axis=1)

    return df[["company_id", "company_name", "broad_sector", "pe_ratio", "pb_ratio", "ev_ebitda",
               "fcf_yield_pct", "five_yr_median_pe", "pe_vs_sector_median_pct", "flag"]].rename(
        columns={"broad_sector": "sector", "pe_ratio": "PE", "pb_ratio": "PB", "ev_ebitda": "EV_EBITDA",
                 "five_yr_median_pe": "5yr_median_PE"})


def export_valuation(df, summary_path="output/valuation_summary.xlsx", flags_path="output/valuation_flags.csv"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    os.makedirs("output", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"
    ws.append(["⚠️ NOTE: P/E, P/B, EV/EBITDA, and market cap figures are SIMULATED data, "
               "not real market prices — for demonstration purposes only."])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(row=1, column=1).font = Font(name="Arial", italic=True, color="C00000")
    ws.append([])
    arial = Font(name="Arial")
    header_font = Font(name="Arial", bold=True)
    caution_fill = PatternFill("solid", fgColor="FFC7CE")
    discount_fill = PatternFill("solid", fgColor="C6EFCE")

    cols = list(df.columns)
    ws.append(cols)
    header_row = 3
    for cell in ws[header_row]:
        cell.font = header_font
    flag_col = cols.index("flag") + 1
    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        ws.append([row[c] for c in cols])
        for cell in ws[r_idx]:
            cell.font = arial
        if row["flag"] == "Caution":
            ws.cell(row=r_idx, column=flag_col).fill = caution_fill
        elif row["flag"] == "Discount":
            ws.cell(row=r_idx, column=flag_col).fill = discount_fill
    wb.save(summary_path)

    flagged = df[df["flag"].isin(["Caution", "Discount"])]
    flagged.to_csv(flags_path, index=False)
    return summary_path, flags_path


def run(db_path=None):
    conn = sqlite3.connect(db_path or DB_PATH)
    df = build_valuation_table(conn)
    if df.empty:
        print("No market_cap data found — cannot build valuation table.")
        conn.close()
        return df
    summary_path, flags_path = export_valuation(df)
    print(f"valuation_summary.xlsx: {len(df)} rows. valuation_flags.csv: "
          f"{len(df[df['flag'].isin(['Caution','Discount'])])} flagged companies.")
    conn.close()
    return df


if __name__ == "__main__":
    run()
