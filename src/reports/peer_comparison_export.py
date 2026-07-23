"""peer_comparison.xlsx exporter — Sprint 3, Day 20.
11 sheets (one per peer group), each with company_id/name + value+percentile for
10 metrics (20 data columns), colour-coded percentile cells, benchmark row highlighted,
and a median summary row.
"""
import os
import sqlite3
import pandas as pd

from src.analytics.peer import METRICS

DB_PATH = os.getenv("DB_PATH", "nifty100.db")


def export_peer_comparison(db_path=None, out_path="output/peer_comparison.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    conn = sqlite3.connect(db_path or DB_PATH)
    pp = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
    pg = pd.read_sql_query("SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn)
    comp = pd.read_sql_query("SELECT id as company_id, company_name FROM companies", conn)
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)
    arial = Font(name="Arial")
    header_font = Font(name="Arial", bold=True)
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    gold = PatternFill("solid", fgColor="FFD966")

    metrics = list(METRICS.keys())
    if pp.empty:
        wb.create_sheet("No Data")
        dirname = os.path.dirname(out_path)
        if dirname:
            os.makedirs(dirname, exist_ok=True)
        wb.save(out_path)
        return out_path

    for group_name, members in pg.groupby("peer_group_name"):
        ws = wb.create_sheet(group_name[:31])
        header = ["company_id", "company_name"] + [f"{m}_value" for m in metrics] + [f"{m}_pctile" for m in metrics]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font

        gdf = pp[pp["peer_group_name"] == group_name]
        company_ids = members["company_id"].tolist()
        row_start = 2
        pctile_col_start = 3 + len(metrics)

        for r_idx, cid in enumerate(company_ids, start=row_start):
            name = comp.loc[comp["company_id"] == cid, "company_name"]
            name = name.iloc[0] if not name.empty else cid
            row = [cid, name]
            values, pctiles = [], []
            for m in metrics:
                rec = gdf[(gdf["company_id"] == cid) & (gdf["metric"] == m)]
                values.append(rec["value"].iloc[0] if not rec.empty else None)
                pctiles.append(rec["percentile_rank"].iloc[0] if not rec.empty else None)
            ws.append(row + values + pctiles)

            is_bench = bool(members.loc[members["company_id"] == cid, "is_benchmark"].iloc[0]) \
                if not members.loc[members["company_id"] == cid].empty else False
            for c_idx in range(1, len(header) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = arial
                if is_bench:
                    cell.fill = gold
            for m_idx, p in enumerate(pctiles):
                if p is None:
                    continue
                cell = ws.cell(row=r_idx, column=pctile_col_start + m_idx)
                if not is_bench:
                    if p >= 0.75:
                        cell.fill = green
                    elif p <= 0.25:
                        cell.fill = red
                    else:
                        cell.fill = yellow

        # median summary row
        last_row = row_start + len(company_ids)
        median_row = ["MEDIAN", ""]
        for m in metrics:
            vals = gdf[gdf["metric"] == m]["value"].dropna()
            median_row.append(vals.median() if not vals.empty else None)
        for m in metrics:
            pv = gdf[gdf["metric"] == m]["percentile_rank"].dropna()
            median_row.append(pv.median() if not pv.empty else None)
        ws.append(median_row)
        for cell in ws[last_row]:
            cell.font = header_font

    dirname = os.path.dirname(out_path)
    if dirname:
        os.makedirs(dirname, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    export_peer_comparison()
