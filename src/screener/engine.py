"""Screener engine — Sprint 3, Day 15-17.
Loads config/screener_config.yaml, builds the latest-year universe, applies filters,
computes composite quality scores (global + sector-relative), and exports
output/screener_output.xlsx with green/red threshold colour-coding.
"""
import os
import sqlite3
import yaml
import pandas as pd

from src.analytics.cagr import revenue_cagr

DB_PATH = os.getenv("DB_PATH", "nifty100.db")
CONFIG_PATH = "config/screener_config.yaml"

FILTER_COL_MAP = {
    "roe_min": "return_on_equity_pct", "de_max": "debt_to_equity", "fcf_min": "free_cash_flow_cr",
    "revenue_cagr_5yr_min": "revenue_cagr_5yr", "pat_cagr_5yr_min": "pat_cagr_5yr",
    "opm_min": "operating_profit_margin_pct", "pe_max": "pe_ratio", "pb_max": "pb_ratio",
    "dividend_yield_min": "dividend_yield_pct", "icr_min": "interest_coverage",
    "market_cap_min": "market_cap_crore", "net_profit_min": "net_profit",
    "eps_cagr_min": "eps_cagr_5yr", "asset_turnover_min": "asset_turnover", "sales_min": "sales",
    "dividend_payout_max": "dividend_payout_ratio_pct", "revenue_cagr_3yr_min": "revenue_cagr_3yr",
}


def load_config(path=CONFIG_PATH):
    with open(path) as f:
        return yaml.safe_load(f)


def _latest_year(df):
    return df.sort_values("year").groupby("company_id").tail(1)


def _revenue_cagr_n(conn, company_id, n):
    hist = pd.read_sql_query("SELECT year, sales FROM profitandloss WHERE company_id=? ORDER BY year",
                              conn, params=(company_id,))
    series = dict(zip(hist["year"], hist["sales"]))
    if not series:
        return None
    latest = max(series.keys())
    val, _flag = revenue_cagr(series, latest, n)
    return val


def _de_declining_yoy(conn, company_id):
    hist = pd.read_sql_query(
        "SELECT year, debt_to_equity FROM financial_ratios WHERE company_id=? ORDER BY year",
        conn, params=(company_id,))
    if len(hist) < 2:
        return False
    last_two = hist.tail(2)["debt_to_equity"].tolist()
    if any(pd.isna(x) for x in last_two):
        return False
    return last_two[1] < last_two[0]


def build_universe(conn):
    """One row per company using each company's latest available year."""
    fr = pd.read_sql_query("SELECT * FROM financial_ratios", conn)
    if fr.empty:
        return fr
    fr = _latest_year(fr)

    comp = pd.read_sql_query("SELECT id as company_id, company_name FROM companies", conn)
    sec = pd.read_sql_query("SELECT company_id, broad_sector FROM sectors", conn)
    pl = pd.read_sql_query("SELECT company_id, year, sales, net_profit FROM profitandloss", conn)
    mc = pd.read_sql_query(
        "SELECT company_id, year, pe_ratio, pb_ratio, dividend_yield_pct, market_cap_crore FROM market_cap", conn)

    df = fr.merge(comp, on="company_id", how="left").merge(sec, on="company_id", how="left")
    df = df.merge(pl, on=["company_id", "year"], how="left")
    if not mc.empty:
        mc_latest = mc.sort_values("year").groupby("company_id").tail(1).drop(columns=["year"])
        df = df.merge(mc_latest, on="company_id", how="left")
    else:
        for c in ("pe_ratio", "pb_ratio", "dividend_yield_pct", "market_cap_crore"):
            df[c] = None

    df["revenue_cagr_3yr"] = df["company_id"].apply(lambda cid: _revenue_cagr_n(conn, cid, 3))
    df["de_declining_yoy"] = df["company_id"].apply(lambda cid: _de_declining_yoy(conn, cid))
    df["fcf_positive_latest"] = df["free_cash_flow_cr"].apply(lambda v: pd.notna(v) and v > 0)
    return df


def _passes(filter_key, value, threshold, sector=None):
    if pd.isna(value):
        return False
    if filter_key == "de_max":
        return sector == "Financials" or value <= threshold  # Financials skip this filter
    if filter_key == "icr_min":
        return True  # None (Debt Free) already excluded upstream by pd.isna check on non-ICR cols;
        # actual ICR None-as-infinity handling is done in apply_filters, not here
    if filter_key.endswith("_min"):
        return value >= threshold
    if filter_key.endswith("_max"):
        return value <= threshold
    if filter_key in ("fcf_positive_latest", "de_declining_yoy"):
        return bool(value) == bool(threshold)
    return True


def apply_filters(df, filters):
    mask = pd.Series(True, index=df.index)
    for key, threshold in filters.items():
        if key == "de_max":
            cond = (df["broad_sector"] == "Financials") | (df["debt_to_equity"] <= threshold)
            mask &= cond.fillna(False)
        elif key == "icr_min":
            cond = df["interest_coverage"].isna() | (df["interest_coverage"] >= threshold)
            mask &= cond
        elif key in ("fcf_positive_latest", "de_declining_yoy"):
            mask &= (df[key] == threshold)
        else:
            col = FILTER_COL_MAP.get(key)
            if col is None or col not in df.columns:
                continue
            if key.endswith("_min"):
                mask &= (df[col] >= threshold)
            elif key.endswith("_max"):
                mask &= (df[col] <= threshold)
    return df[mask]


def winsor_scale(series, invert=False):
    vals_sorted = sorted(v for v in series if pd.notna(v))
    def _s(v):
        if pd.isna(v) or not vals_sorted:
            return 50.0
        p10 = vals_sorted[int(0.10 * (len(vals_sorted) - 1))]
        p90 = vals_sorted[int(0.90 * (len(vals_sorted) - 1))]
        if p90 == p10:
            return 50.0
        x = max(p10, min(p90, v))
        s = (x - p10) / (p90 - p10) * 100
        return 100 - s if invert else s
    return series.apply(_s)


def _grouped_scale(df, col, group_col=None, invert=False):
    if group_col and group_col in df.columns:
        return df.groupby(group_col)[col].transform(lambda s: winsor_scale(s, invert))
    return winsor_scale(df[col], invert)


def compute_composite_scores(df, sector_relative=False):
    """0-100 composite score: 35% profitability + 30% cash quality + 20% growth + 15% leverage.
    NOTE: 'cash quality' uses free_cash_flow_cr magnitude as a proxy for the spec's
    'FCF CAGR' component, since a dedicated FCF CAGR series isn't stored separately —
    documented assumption, see README."""
    grp = "broad_sector" if sector_relative else None

    roe_s = _grouped_scale(df, "return_on_equity_pct", grp)
    roce_s = _grouped_scale(df, "return_on_capital_employed_pct", grp)
    npm_s = _grouped_scale(df, "net_profit_margin_pct", grp)
    fcf_s = _grouped_scale(df, "free_cash_flow_cr", grp)
    cfo_pat_s = _grouped_scale(df, "cfo_quality_score", grp)
    fcf_pos_s = df["free_cash_flow_cr"].apply(lambda v: 100.0 if pd.notna(v) and v > 0 else 0.0)
    rev_cagr_s = _grouped_scale(df, "revenue_cagr_5yr", grp)
    pat_cagr_s = _grouped_scale(df, "pat_cagr_5yr", grp)

    de_proxy = df["debt_to_equity"].apply(lambda v: None if pd.isna(v) or v == float("inf") else -v)
    df_tmp = df.copy()
    df_tmp["_de_proxy"] = de_proxy
    de_s = _grouped_scale(df_tmp, "_de_proxy", grp)

    icr_proxy = df["interest_coverage"].apply(lambda v: 999.0 if pd.isna(v) else v)  # Debt Free -> best
    df_tmp["_icr_proxy"] = icr_proxy
    icr_s = _grouped_scale(df_tmp, "_icr_proxy", grp)

    profitability = 0.15 * roe_s + 0.10 * roce_s + 0.10 * npm_s
    cash_quality = 0.15 * fcf_s + 0.10 * cfo_pat_s + 0.05 * fcf_pos_s
    growth = 0.10 * rev_cagr_s + 0.10 * pat_cagr_s
    leverage = 0.10 * de_s + 0.05 * icr_s
    return (profitability + cash_quality + growth + leverage).round(1)


def run_preset(df, preset_key, config):
    preset = config["presets"][preset_key]
    result = apply_filters(df, preset["filters"])
    return result.sort_values("composite_score", ascending=False)


def run_all_presets(df, config):
    return {k: run_preset(df, k, config) for k in config["presets"]}


def export_screener_output(results, config, path="output/screener_output.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    arial = Font(name="Arial")
    header_font = Font(name="Arial", bold=True)

    display_cols = ["company_id", "company_name", "broad_sector", "composite_score",
                     "return_on_equity_pct", "return_on_capital_employed_pct",
                     "net_profit_margin_pct", "debt_to_equity", "interest_coverage",
                     "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
                     "pe_ratio", "pb_ratio", "dividend_yield_pct", "dividend_payout_ratio_pct",
                     "asset_turnover", "sales", "net_profit", "market_cap_crore"]

    for preset_key, df in results.items():
        label = config["presets"][preset_key]["label"]
        ws = wb.create_sheet(label[:31])
        cols = [c for c in display_cols if c in df.columns]
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
        for _, row in df.iterrows():
            ws.append([row.get(c) for c in cols])
        for r_row in ws.iter_rows(min_row=2):
            for cell in r_row:
                cell.font = arial

        filters = config["presets"][preset_key]["filters"]
        for filter_key, threshold in filters.items():
            col_name = FILTER_COL_MAP.get(filter_key)
            if col_name not in cols:
                continue
            col_idx = cols.index(col_name) + 1
            for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
                sector = row.get("broad_sector")
                val = row.get(col_name)
                if filter_key == "de_max" and sector == "Financials":
                    continue  # not evaluated for this filter, leave uncoloured
                passed = _passes(filter_key, val, threshold, sector)
                ws.cell(row=r_idx, column=col_idx).fill = green if passed else red

    dirname = os.path.dirname(path)
    if dirname:
        os.makedirs(dirname, exist_ok=True)
    wb.save(path)
    return path


def run(db_path=None, config_path=CONFIG_PATH, out_path="output/screener_output.xlsx"):
    conn = sqlite3.connect(db_path or DB_PATH)
    config = load_config(config_path)
    universe = build_universe(conn)
    universe["composite_score"] = compute_composite_scores(universe, sector_relative=False)
    results = run_all_presets(universe, config)
    for k, df in results.items():
        n = len(df)
        print(f"{config['presets'][k]['label']}: {n} companies")
    path = export_screener_output(results, config, out_path)
    conn.close()
    return results, path


if __name__ == "__main__":
    run()
